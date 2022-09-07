from __future__ import absolute_import
from es_gui.apps.data_manager.utils import check_connection_settings
from es_gui.proving_grounds.charts import RateScheduleChart
from es_gui.apps.data_manager.data_manager import DataManagerException, DATA_HOME
from es_gui.resources.widgets.common import InputError, WarningPopup, ConnectionErrorPopup, MyPopup, RecycleViewRow, FADEIN_DUR, LoadingModalView, PALETTE, rgba_to_fraction, fade_in_animation, DataGovAPIhelp, ParameterRow

import os
import io
import calendar
import datetime
import logging
import datetime as dt
import webbrowser
import json
import csv
from codecs import iterdecode
from collections import defaultdict
from openpyxl import load_workbook

import requests
import pandas as pd
import numpy as np
from kivy.animation import Animation
from kivy.app import App
from kivy.clock import Clock, mainthread
from kivy.uix.screenmanager import ScreenManager, Screen, SlideTransition
from kivy.uix.modalview import ModalView
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.gridlayout import GridLayout
from kivy.properties import ObjectProperty, NumericProperty, BooleanProperty, StringProperty, DictProperty

from kivy.lang import Builder

import urllib3
urllib3.disable_warnings(
    requests.packages.urllib3.exceptions.InsecureRequestWarning)


MAX_WHILE_ATTEMPTS = 7

URL_FACT = "https://api.epa.gov/FACT/1.0/emissions/hourlyData/csv/"
URL_FACT_Facilities = "https://api.epa.gov/FACT/1.0/facilities?api_key="

URL_PVWATTS = "https://developer.nrel.gov/api/pvwatts/v6.json?"

PPNC_URL = 'https://www.epa.gov/sites/default/files/2021-05/power_plants_and_communities.xlsx'

STATIC_HOME = 'es_gui/apps/data_manager/_static'


class PowerPlantSearchScreen(Screen):
    """DataManager screen for searching for Power Plant Dispatch and pollution  Data through FACT."""

    api_key = StringProperty('')

    def __init__(self, **kwargs):
        super(PowerPlantSearchScreen, self).__init__(**kwargs)

        PowerPlantSaveNameTextInput.host_screen = self

    def on_pre_enter(self):
        if not self.param_widget.children:
            self.param_widget.build()

    def on_enter(self):

        def _ref_link(text, ref):
            return '[ref={0}][color=003359][u]{1}[/u][/color][/ref]'.format(ref, text)

        def _go_to_webpage(instance, value):
            if value == 'fact-api':
                webbrowser.open(
                    'https://www.epa.gov/airmarkets/field-audit-checklist-tool-fact')
            elif value == 'pv-watts':
                webbrowser.open('https://pvwatts.nrel.gov/')
            elif value == 'ppnc-spreadsheet':
                webbrowser.open(
                    'https://www.epa.gov/airmarkets/power-plants-and-neighboring-communities')
            elif value == 'cobra-api':
                webbrowser.open('https://www.epa.gov/cobra')
            elif value == 'justice40':
                webbrowser.open('https://www.epa.gov/cobra')

        ab = self.manager.nav_bar
        ab.set_title('Data Manager: Power Plant Dispatch and Pollution  Data')
        text1 = 'This program retrieves data from: \n 1) the EPA EPA Field Audit Checklist Tool ({FACT}), \n '.format(
            FACT=_ref_link('FACT', 'fact-api'))
        text2 = '2) NREL\'s {PVWatts}, and \n 3) The EPA Power Plant and Neighboring Communities ({PPNC}) spreadsheet \n'.format(
            PVWatts=_ref_link('PVWatts', 'pv-watts'), PPNC=_ref_link('PPNC', 'ppnc-spreadsheet'))
        text3 = 'It then uses the power plant pollution data to estimate county-level health impacts through the EPA CO-Benefits Risk Assessment ({COBRA}) tool.'.format(
            COBRA=_ref_link('COBRA', 'cobra-api'))
        text4 = 'Last, it multiplies the health impacts by the demography from {Justice40} in each county to determine the distributional equitably.'.format(
            Justice40=_ref_link('Justice40', 'justice40'))
        self.explainer_text.text = text1 + text2 + text3 + text4

        self.explainer_text.bind(on_ref_press=_go_to_webpage)

    def open_api_key_help(self):
        """Opens the API key help ModalView."""
        open_ei_help_view = DataGovAPIhelp()
        open_ei_help_view.open()

    def reset_screen(self):
        """Resets the screen to its initial state."""
        pass

    def _validate_inputs(self):
        """Validates the search parameters."""
        # Check if an API key has been provided.
        api_key = self.api_key_input.text
        # oris locationid year quarter
        if not api_key:
            raise (InputError('Please enter an API key.'))

        power_plant_params = self.param_widget.get_inputs()

        return api_key, power_plant_params

    def get_inputs(self):
        """Retrieves the search inputs and validates them."""
        api_key, power_plant_params = self._validate_inputs()

        # Fixed values.
        power_plant_params['api_key'] = api_key
        #power_plant_params['location_id'] = '1'
        #power_plant_params['quarter'] = '1'

        return api_key, power_plant_params

    def execute_facilities_query(self):
        """Executes the FACT query using the given parameters."""
        try:
            api_key, power_plant_params = self.get_inputs()
            self.api_key = api_key
        except ValueError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        except InputError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        else:
            # Reset screen.
            self.reset_screen()
            self.facilities_button.disabled = True

            try:
                self._facilities_query_api()
            except requests.ConnectionError:
                popup = ConnectionErrorPopup()
                popup.popup_text.text = 'There was an issue connecting to the API. Check your connection settings and try again.'
                popup.open()

    def _facilities_query_api(self):
        """Uses FACT API to query for a PV profile."""
        ssl_verify, proxy_settings = check_connection_settings()

        # Form query.
        api_query = URL_FACT_Facilities + self.api_key

        try:
            with requests.Session() as req:
                http_request = req.get(api_query,
                                       proxies=proxy_settings,
                                       timeout=10,
                                       verify=ssl_verify,
                                       stream=True)
                if http_request.status_code != requests.codes.ok:
                    http_request.raise_for_status()
        except requests.HTTPError as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except requests.exceptions.ProxyError:
            logging.error('PowerPlantDM: Could not connect to proxy.')
            raise requests.ConnectionError
        except requests.ConnectionError as e:
            logging.error(
                'PowerPlantDM: Failed to establish a connection to the host server.')
            raise requests.ConnectionError
        except requests.Timeout as e:
            logging.error('PowerPlantDM: The connection timed out.')
            raise requests.ConnectionError
        except requests.RequestException as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except Exception as e:
            # Something else went wrong.
            logging.error(
                'PowerPlantDM: An unexpected error has occurred. ({0})'.format(repr(e)))
            raise requests.ConnectionError
        else:
            request_content = http_request.json()

            outname = 'facilities'
            # Strip non-alphanumeric chars from given name for filename.
            delchars = ''.join(c for c in map(
                chr, range(256)) if not c.isalnum())
            outname = outname.translate({ord(i): None for i in delchars})

            # Save facilities list
            destination_file = os.path.join(STATIC_HOME, outname + '.json')

            with open(destination_file, 'w') as outfile:
                json.dump(request_content, outfile)

                logging.info(
                    'PowerPlantDM: Facilities list successfully saved.')

                popup = WarningPopup()
                popup.title = 'Success!'
                popup.popup_text.text = 'Facilities list successfully saved.'
                popup.open()
        finally:
            self.facilities_button.disabled = False

    def execute_PPNC_query(self):
        """Executes a download of the Power Plants and Neighboring Communities spreadsheet."""
        try:
            api_key, power_plant_params = self.get_inputs()
            self.api_key = api_key
        except ValueError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        except InputError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        else:
            # Reset screen.
            self.reset_screen()
            self.ppnc_button.disabled = True

            try:
                self._PPNC_query_api()
            except requests.ConnectionError:
                popup = ConnectionErrorPopup()
                popup.popup_text.text = 'There was an issue connecting to the API. Check your connection settings and try again.'
                popup.open()

    def _PPNC_query_api(self):
        """Download of the Power Plants and Neighboring Communities spreadsheet."""
        ssl_verify, proxy_settings = check_connection_settings()

        try:
            with requests.Session() as req:
                http_request_ppc = req.get(PPNC_URL,
                                           proxies=proxy_settings,
                                           timeout=10,
                                           verify=ssl_verify,
                                           stream=True)
                if http_request_ppc.status_code != requests.codes.ok:
                    http_request_ppc.raise_for_status()
        except requests.HTTPError as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except requests.exceptions.ProxyError:
            logging.error('PowerPlantDM: Could not connect to proxy.')
            raise requests.ConnectionError
        except requests.ConnectionError as e:
            logging.error(
                'PowerPlantDM: Failed to establish a connection to the host server.')
            raise requests.ConnectionError
        except requests.Timeout as e:
            logging.error('PowerPlantDM: The connection timed out.')
            raise requests.ConnectionError
        except requests.RequestException as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except Exception as e:
            # Something else went wrong.
            logging.error(
                'PowerPlantDM: An unexpected error has occurred. ({0})'.format(repr(e)))
            raise requests.ConnectionError
        else:
            # Save power_plants_and_communities.xlsx
            destination_file_ppc = os.path.join(
                STATIC_HOME, 'power_plants_and_communities.xlsx')
            with open(destination_file_ppc, 'wb') as outfile:
                outfile.write(http_request_ppc.content)

                logging.info(
                    'PowerPlantDM: Power Plants and Neighboring Communities spreadsheet successfully saved.')

                popup = WarningPopup()
                popup.title = 'Success!'
                popup.popup_text.text = 'Power Plants and Neighboring Communities spreadsheet successfully saved.'
                popup.open()
        finally:
            self.ppnc_button.disabled = False

    def execute_query(self):
        """Executes the FACT query using the given parameters."""
        try:
            api_key, power_plant_params = self.get_inputs()
            self.api_key = api_key
        except ValueError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        except InputError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        else:
            # Reset screen.
            self.reset_screen()
            self.save_button.disabled = True

            try:
                request_content = self._query_api(power_plant_params)
            except requests.ConnectionError:
                popup = ConnectionErrorPopup()
                popup.popup_text.text = 'There was an issue connecting to the API. Check your connection settings and try again.'
                popup.open()

    def _query_api(self, power_plant_params):
        """Uses FACT API to query for a normalized PV profile and powerplant data."""
        ssl_verify, proxy_settings = check_connection_settings()

        http_request = defaultdict(lambda: defaultdict(dict))

        if not self.save_name_field.text:
            popup = WarningPopup()
            popup.popup_text.text = 'Please specify a name to save the power plant as.'
            popup.open()
            return
        else:
            outname = self.save_name_field.text

        # pull the site location from the facilities data file
        facilities_file = os.path.join(STATIC_HOME, 'facilities.json')
        unit_list = []
        with open(facilities_file, 'r') as f:
            facilities_data = json.load(f)
        for faci in facilities_data['data']:
            if faci['orisCode'] == int(power_plant_params['Plant_id']):
                name = faci['name']
                latitude = faci['geographicLocation']['latitude']
                longitude = faci['geographicLocation']['longitude']
                nameplateCapacity = 0
                if power_plant_params['location_id'] == '0':
                    for unit in faci['units']:
                        unit_list.append(unit['unitId'])

        # Set up an api queary to get PV Watts solar data from the same location as the facility
        pv_api_query = URL_PVWATTS
        query_segs = []
        pv_params = ([('azimuth', 180),
                      ('losses', 14),
                      ('system_capacity', 1000),
                      ('lon', longitude),
                      ('lat', latitude),
                      ('module_type', 0),
                      ('array_type', 0),
                      ('dataset', 'tmy3'),
                      ('radius', 0),
                      ('timeframe', 'hourly'),
                      ('api_key', self.api_key),
                      ('tilt', latitude)])

        for k, v in pv_params:
            query_segs.append('{key}={value}'.format(key=k, value=v))

        pv_api_query += '&'.join(query_segs)

        try:
            with requests.Session() as req:
                logging.info(
                    'PowerPlantDM: sending api request for power plant data')
                for quarter in range(4):
                    if power_plant_params['location_id'] == '0':
                        for unit in unit_list:
                            # Form query.
                            api_query = URL_FACT + (power_plant_params['Plant_id'] + '/' +
                                                    unit + '/' +
                                                    power_plant_params['year'] + '/' +
                                                    str(quarter+1) +
                                                    '?api_key=' + self.api_key)
                            # print(api_query)
                            http_request[unit][quarter] = req.get(api_query,
                                                                  proxies=proxy_settings,
                                                                  timeout=10,
                                                                  verify=ssl_verify,
                                                                  stream=True)

                            if http_request[unit][quarter].status_code != requests.codes.ok:
                                http_request[unit][quarter].raise_for_status()
                                break

                    else:
                        api_query = URL_FACT + (power_plant_params['Plant_id'] + '/' +
                                                power_plant_params['location_id'] + '/' +
                                                power_plant_params['year'] + '/' +
                                                str(quarter+1) +
                                                '?api_key=' + self.api_key)
                        # print(api_query)
                        http_request[quarter] = req.get(api_query,
                                                        proxies=proxy_settings,
                                                        timeout=10,
                                                        verify=ssl_verify,
                                                        stream=True)

                        if http_request[quarter].status_code != requests.codes.ok:
                            http_request[quarter].raise_for_status()
                            break

                logging.info(
                    'PowerPlantDM: sending api request for pv data at power plant site')
                http_request_pv = req.get(pv_api_query,
                                          proxies=proxy_settings,
                                          timeout=10,
                                          verify=ssl_verify,
                                          stream=True)
                if http_request_pv.status_code != requests.codes.ok:
                    http_request_pv.raise_for_status()
        except requests.HTTPError as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except requests.exceptions.ProxyError:
            logging.error('PowerPlantDM: Could not connect to proxy.')
            raise requests.ConnectionError
        except requests.ConnectionError as e:
            logging.error(
                'PowerPlantDM: Failed to establish a connection to the host server.')
            raise requests.ConnectionError
        except requests.Timeout as e:
            logging.error('PowerPlantDM: The connection timed out.')
            raise requests.ConnectionError
        except requests.RequestException as e:
            logging.error('PowerPlantDM: {0}'.format(repr(e)))
            raise requests.ConnectionError
        except Exception as e:
            # Something else went wrong.
            logging.error(
                'PowerPlantDM: An unexpected error has occurred. ({0})'.format(repr(e)))
            raise requests.ConnectionError
        else:
            request_content = dict()
            request_content['Name'] = name
            request_content['Oris_id'] = int(power_plant_params['Plant_id'])
            request_content['lat'] = latitude
            request_content['lon'] = longitude
            request_content['year'] = int(power_plant_params['year'])

            # cross referance with power plant comunities spreadsheet for pollution  numbers
            logging.info(
                'PowerPlantDM: loading large power plant comunities spreadsheet')
            ppc_data_file = os.path.join(
                STATIC_HOME, 'power_plants_and_communities.xlsx')
            workbook = load_workbook(filename=ppc_data_file)
            Plants_spreadsheet = workbook.active
            # column 6 is the DOE/EIA ORIS plant or facility code
            i = 1
            row_number = 0
            for value in Plants_spreadsheet.iter_rows(min_col=6, max_col=6, values_only=True):
                if value[0] == request_content['Oris_id']:
                    row_number = i
                    break
                i += 1

            if row_number == 0:
                logging.info(
                    'PowerPlantDM: Error: could not find Oris_id in power plant comunities spreadsheet')
            else:
                # Plant nameplate capacity (MW) 29
                request_content['NameplateCapacity'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=29).value)
                # Plant FIPS state code 18
                request_content['Plant_FIPS_state_code'] = Plants_spreadsheet.cell(
                    row=row_number, column=18).value
                # Plant FIPS county code 19
                request_content['Plant_FIPS_county_code'] = Plants_spreadsheet.cell(
                    row=row_number, column=19).value
                # Plant capacity factor 28
                request_content['CapacityFactor'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=28).value)
                # Plant primary coal/oil/gas/ other fossil fuel category 26
                request_content['FuelCategory'] = Plants_spreadsheet.cell(
                    row=row_number, column=26).value
                # Plant associated ISO/RTO Territory (null if N/A) 17
                request_content['ISORTOTerritory'] = Plants_spreadsheet.cell(
                    row=row_number, column=17).value
                # Total population within 3 miles (ACS2018) 150
                request_content['population'] = int(
                    Plants_spreadsheet.cell(row=row_number, column=150).value)
                # PM 2.5 Emssions (tons)            318
                request_content['PPNC_PM25_emissions'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=318).value)
                # Plant annual NOx emissions (tons) 43
                request_content['PPNC_NOx_emissions'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=43).value)
                # Plant annual SO2 emissions (tons) 45
                request_content['PPNC_SO2_emissions'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=45).value)
                # Plant annual CO2 emissions (tons) 46
                request_content['PPNC_CO2_emissions'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=46).value)
                # Plant annual net generation (MWh) 41
                request_content['PPNC_net_generation'] = float(
                    Plants_spreadsheet.cell(row=row_number, column=46).value)

                request_content['PM25_emissions_rate'] = request_content['PPNC_PM25_emissions'] / \
                    request_content['PPNC_net_generation']
                request_content['NOx_emissions_rate'] = request_content['PPNC_NOx_emissions'] / \
                    request_content['PPNC_net_generation']
                request_content['SO2_emissions_rate'] = request_content['PPNC_SO2_emissions'] / \
                    request_content['PPNC_net_generation']
                request_content['CO2_emissions_rate'] = request_content['PPNC_CO2_emissions'] / \
                    request_content['PPNC_net_generation']

            # log plant dispatch and pv resource avalibuility into request_content directory
            request_content_pv = http_request_pv.json()
            pv_profile = request_content_pv['outputs']['ac']
            # note system_capacity is in kW whereas the output ac is in watts
            # PV profile is normalized to the range [0,1]
            pv_profile = [pv_profile[i]/(float(request_content_pv['inputs']
                                         ['system_capacity'])*1000) for i in range(len(pv_profile))]
            request_content['plant_dispatch'] = []
            request_content['pv'] = pv_profile
            bace_length = 0
            if power_plant_params['location_id'] == '0':
                for quarter in range(4):
                    i = 0
                    for unit in unit_list:
                        http_request_content = http_request[unit][quarter].content.decode(
                            'utf-8')
                        cr = csv.reader(
                            http_request_content.splitlines(), delimiter=',')

                        data_list = list(cr)
                        j = 0
                        for row in data_list:
                            if i == 0:  # for the first unit in the list add a new element to each list
                                if j >= 1:  # skip the row of data lables
                                    if row[2].isnumeric():
                                        #print(row[2] + str(row[2].isnumeric()))
                                        power = float(row[2])
                                        #print('power : ' + str(power))
                                        request_content['plant_dispatch'].append(
                                            power)
                                    else:
                                        request_content['plant_dispatch'].append(
                                            0.0)
                            else:  # after that add each unit's plant_dispatch onto the total for the facility
                                if j >= 1:  # skip the row of data lables
                                    if row[2].isnumeric():
                                        power = float(row[2])
                                        #print('power : ' + str(power))
                                        request_content['plant_dispatch'][bace_length+j -
                                                                          1] = request_content['plant_dispatch'][bace_length+j-1] + power
                            j += 1
                        i += 1
                    bace_length = len(request_content['plant_dispatch'])
            else:
                for quarter in range(4):
                    http_request_content = http_request[quarter].content.decode(
                        'utf-8')
                    cr = csv.reader(
                        http_request_content.splitlines(), delimiter=',')

                    data_list = list(cr)
                    j = 0
                    for row in data_list:
                        if j >= 1:  # skip the row of data lables
                            try:
                                request_content['plant_dispatch'].append(
                                    float(row[2]))
                            except:
                                request_content['plant_dispatch'].append(0.0)
                        j += 1
            request_content['plant_total_MWh'] = sum(
                request_content['plant_dispatch'])

            request_content['PM25_emissions'] = request_content['PM25_emissions_rate'] * \
                request_content['plant_total_MWh']
            request_content['NOx_emissions'] = request_content['NOx_emissions_rate'] * \
                request_content['plant_total_MWh']
            request_content['SO2_emissions'] = request_content['SO2_emissions_rate'] * \
                request_content['plant_total_MWh']
            request_content['CO2_emissions'] = request_content['CO2_emissions_rate'] * \
                request_content['plant_total_MWh']

            # Queary the COBRA API to calculate health impacts and values in each county
            logging.info(
                'PowerPlantDM: sending api request for EPA COBRA tool')
            api_query_token = "https://cobraapi.app.cloud.gov/api/token"
            api_query_SummarizedControlEmissions = 'https://cobraapi.app.cloud.gov/api/SummarizedControlEmissions?'
            api_post_EmissionsUpdate = 'https://cobraapi.app.cloud.gov/api/EmissionsUpdate'
            api_get_Result = 'https://cobraapi.app.cloud.gov/api/Result/'

            fipscodes = request_content['Plant_FIPS_state_code'] + \
                request_content['Plant_FIPS_county_code']
            # tier 1 applys to power plants due to the elivation and despersion of emmision
            tiers = "1"

            with requests.Session() as req:
                # request a simulation token
                http_request_token = req.get(api_query_token,
                                             proxies=proxy_settings,
                                             timeout=30,
                                             verify=ssl_verify,
                                             stream=True)
                request_content_token = http_request_token.json()
                token = request_content_token['value']

                # get the baseline emisions for the state / county where the power plant is located
                api_query_SummarizedControlEmissions = api_query_SummarizedControlEmissions + \
                    'token=' + token + '&fipscodes=' + fipscodes + '&tiers=' + tiers
                http_request_BE = req.get(api_query_SummarizedControlEmissions,
                                          proxies=proxy_settings,
                                          timeout=10,
                                          verify=ssl_verify,
                                          stream=True)
                request_content_json = http_request_BE.json()
                if request_content_json == {'baseline': [], 'control': []}:
                    logging.info(
                        'PowerPlantDM: ERROR tier 1 polution in the selected county is empty in COBRA\'s data ')
                    request_content['COBRA_results'] = 'ERROR tier 1 polution in the selected county is empty in COBRA\'s data'
                else:
                    payload = request_content_json['baseline'][0]
                    print('Baseline NOX  : ' + str(payload['NO2']))
                    print('Baseline SO2  : ' + str(payload['SO2']))
                    print('Baseline PM25 : ' + str(payload['PM25']))
                    print('Peaker NOX  : ' +
                          str(request_content['NOx_emissions']))
                    print('Peaker SO2  : ' +
                          str(request_content['SO2_emissions']))
                    print('Peaker PM25 : ' +
                          str(request_content['PM25_emissions']))
                    payload['NO2'] = max(
                        [0, payload['NO2'] - request_content['NOx_emissions']])
                    payload['SO2'] = max(
                        [0, payload['SO2'] - request_content['SO2_emissions']])
                    payload['PM25'] = max(
                        [0, payload['PM25'] - request_content['PM25_emissions']])
                    print('Final NOX  : ' + str(payload['NO2']))
                    print('Final SO2  : ' + str(payload['SO2']))
                    print('Final PM25 : ' + str(payload['PM25']))

                    # tier 1 applys to power plants due to the elivation and despersion of emmisions
                    request_body = {
                        "spec": {
                            "fipscodes": [
                                fipscodes,
                            ],
                            "tiers": "1",
                            "token": token
                        },
                        "payload": payload
                    }

                    http_request_post_EmissionsUpdate = req.post(api_post_EmissionsUpdate,
                                                                 json=request_body,
                                                                 proxies=proxy_settings,
                                                                 timeout=30,
                                                                 verify=ssl_verify,
                                                                 stream=True)

                    api_get_Result = api_get_Result + token + '/00'
                    http_request_post_Result = req.get(api_get_Result,
                                                       proxies=proxy_settings,
                                                       timeout=30,
                                                       verify=ssl_verify,
                                                       stream=True)

                    request_content['COBRA_results'] = http_request_post_Result.json(
                    )

            # Determin how much of the health impacts of the peaker
            # pollution  is accrued by 1) disadvantage communities and 2) by
            # people with low income (<200% poverty)
            logging.info(
                'PowerPlantDM: loading justice40 community data to calculate pollution impact equity')
            disadvantaged_population_file = os.path.join(
                STATIC_HOME, 'disadvantaged_pop_by_county_2010.csv')
            pop = {}
            dis = {}
            low = {}
            total_pop = 0
            total_dis = 0
            total_low = 0
            header = True
            with open(disadvantaged_population_file, newline='') as csvfile:
                csv_file = csv.reader(csvfile, delimiter=',')
                for row in csv_file:
                    if not header:
                        ID = row[0]
                        pop[ID] = float(row[1])
                        dis[ID] = float(row[2])
                        low[ID] = float(row[3])
                        total_pop = total_pop + pop[ID]
                        total_dis = total_dis + pop[ID]*dis[ID]
                        total_low = total_low + pop[ID]*low[ID]
                    header = False

            logging.info('PowerPlantDM: calculating pollution impact equity')
            l_value_to_dis_per_county = []
            h_value_to_dis_per_county = []
            l_value_to_low_per_county = []
            h_value_to_low_per_county = []

            l_total_value = request_content['COBRA_results']['Summary']['TotalHealthBenefitsValue_low']
            h_total_value = request_content['COBRA_results']['Summary']['TotalHealthBenefitsValue_high']

            ben_frac = {}
            max_frac = 0
            for impact in request_content['COBRA_results']['Impacts']:
                FIPS = impact['FIPS']
                # Shannon County, SD (FIPS code = 46113) was renamed Oglala Lakota County and assigned anew FIPS code (46102) effective in 2014.
                if FIPS == '46102':
                    FIPS = '46113'
                LV = impact['C__Total_Health_Benefits_Low_Value']
                HV = impact['C__Total_Health_Benefits_High_Value']
                ben_frac[FIPS] = HV/h_total_value
                if ben_frac[FIPS] > max_frac:
                    max_frac = ben_frac[FIPS]
                l_value_to_dis_per_county.append(dis[FIPS]*LV)
                h_value_to_dis_per_county.append(dis[FIPS]*HV)
                l_value_to_low_per_county.append(low[FIPS]*LV)
                h_value_to_low_per_county.append(low[FIPS]*HV)

            #print('low estemate of total health benfits  : $' + str(l_total_value))
            #print('high estemate of total health benfits : $' + str(h_total_value))
            #print('% of total population in disadvantaged comunities :' + str(total_dis/total_pop))
            #print('% of total health benfits to disadvantaged comunities :' + str(sum(h_value_to_dis_per_county)/h_total_value))
            #print('% of total population that is low income (<200% of poverty):' + str(total_low/total_pop))
            #print('% of total health benfits to low income (<200% of poverty):' + str(sum(h_value_to_low_per_county)/h_total_value))
            request_content['health_impact_equity'] = {}
            request_content['health_impact_equity']['total_population'] = total_pop
            request_content['health_impact_equity']['total_disadvantaged_population'] = total_dis
            request_content['health_impact_equity']['total_low_income_population'] = total_low
            request_content['health_impact_equity']['disadvantaged_population_fraction'] = total_dis/total_pop
            request_content['health_impact_equity']['low_income_population_fraction'] = total_low/total_pop
            request_content['health_impact_equity']['total_impact_on_disadvantaged_population_low'] = sum(
                l_value_to_dis_per_county)
            request_content['health_impact_equity']['total_impact_on_disadvantaged_population_high'] = sum(
                h_value_to_dis_per_county)
            request_content['health_impact_equity']['total_impact_on_low_income_population_low'] = sum(
                l_value_to_low_per_county)
            request_content['health_impact_equity']['total_impact_on_low_income_population_high'] = sum(
                h_value_to_low_per_county)
            request_content['health_impact_equity']['impact_on_disadvantaged_population_fraction'] = sum(
                h_value_to_dis_per_county)/h_total_value
            request_content['health_impact_equity']['impact_on_low_income_population_fraction'] = sum(
                h_value_to_low_per_county)/h_total_value

            # Strip non-alphanumeric chars from given name for filename.
            delchars = ''.join(c for c in map(
                chr, range(256)) if not c.isalnum())
            outname = outname.translate({ord(i): None for i in delchars})

            # Save.
            destination_dir = os.path.join(DATA_HOME, 'power_plant')
            os.makedirs(destination_dir, exist_ok=True)
            destination_file = os.path.join(destination_dir, outname + '.json')

            logging.info(
                'PowerPlantDM: saving power plant data file to: ' + destination_file)
            if not os.path.exists(destination_file):
                with open(destination_file, 'w') as outfile:
                    json.dump(request_content, outfile)

                logging.info(
                    'PowerPlantDM: power plant data successfully saved.')

                popup = WarningPopup()
                popup.title = 'Success!'
                popup.popup_text.text = 'Power plant data successfully saved.'
                if payload['NO2'] == 0:
                    popup.popup_text.text += '\n Warning: baseline NOX emissions < powerplant NOX emissions, set to 0 tons'
                if payload['SO2'] == 0:
                    popup.popup_text.text += '\n Warning: baseline SO2 emissions < powerplant SO2 emissions, set to 0 tons'
                if payload['PM25'] == 0:
                    popup.popup_text.text += '\n Warning: baseline PM2.5 emissions < powerplant PM2.5 emissions, set to 0 tons'
                popup.open()
            else:
                # File already exists with same name.
                popup = WarningPopup()
                popup.open()
        finally:
            self.save_button.disabled = False


class PowerPlantSaveNameTextInput(TextInput):
    """TextInput field for entering the save name in the Power Plant Data Manager."""
    host_screen = None

    def keyboard_on_key_down(self, window, keycode, text, modifiers):
        key, key_str = keycode

        if key_str in ('enter', 'numpadenter'):
            self.host_screen.execute_query()
        else:
            super(PowerPlantSaveNameTextInput, self).keyboard_on_key_down(
                window, keycode, text, modifiers)

        return True


class PowerPlantSearchParameterWidget(GridLayout):
    """Grid layout containing rows of parameter adjustment widgets."""

    def build(self):
        # Build the widget by creating a row for each parameter.
        data_manager = App.get_running_app().data_manager
        MODEL_PARAMS = data_manager.get_power_plant_model_params()

        for param in MODEL_PARAMS:
            row = ParameterRow(desc=param)
            self.add_widget(row)
            setattr(self, param['attr name'], row)

    def _validate_inputs(self):
        param_set = {}

        for row in self.children:
            attr_name = row.desc['attr name']

            if not row.text_input.text:
                attr_val = row.text_input.hint_text
            else:
                attr_val = row.text_input.text

            if attr_val:
                param_set[attr_name] = attr_val
            else:
                # Skip if values is None.
                continue

        return param_set

    def get_inputs(self):
        try:
            params = self._validate_inputs()
        except InputError as e:
            popup = WarningPopup()
            popup.popup_text.text = str(e)
            popup.open()
        else:
            return params
